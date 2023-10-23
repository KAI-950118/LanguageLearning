import dask.dataframe as dd
import numpy as np
import openpyxl
import pandas as pd

import Azure_TTS as TTS
import ChatGPT_API as GPT


## GetPath
def GetExcelPath():
    with open('ExcelPath.txt', 'r', encoding='utf-8') as file:
        content = file.read()
    return content


## Cover
def CoverRead(path='', sheet='Cover'):
    df = pd.read_excel(path, sheet)
    result_dict = df.set_index('Item').to_dict()['Num']
    return result_dict


def CoverUpdate(CoverDict, path='', sheet='Cover'):
    df = pd.DataFrame(list(CoverDict.items()), columns=['Item', 'Num'])
    # df.to_excel(path, sheet, index=False)
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]

    for key, value in CoverDict.items():
        for row_num in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_num, column=1).value == key:
                sheet.cell(row=row_num, column=2, value=value)
                break
    workbook.save(path)


## Vocabulary
def VocaUpdate(Vocabulary, path='', sheet='Vocabulary'):
    # 先確定有沒有重複
    Replace = VocaCheck(Vocabulary, path=path)
    if Replace == False:
        # 知道上次最後一行
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet]
        last_row = sheet.max_row
        LastId = last_row + 1
        # 更新內容
        try:
            Vocabulary = Vocabulary.lower()
        except:
            Vocabulary = Vocabulary
        data_list = [LastId, Vocabulary, 0]
        for col_num, value in enumerate(data_list, start=1):
            sheet.cell(row=LastId, column=col_num, value=value)
        workbook.save(path)


def VocaCheck(Vocabulary, path='', sheet='Vocabulary'):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]
    word_list = [cell.value.lower() for cell in sheet['B'] if cell.value]
    if Vocabulary.lower() in word_list:
        # print(f"{Vocabulary} is exist")
        return True
    else:
        # print(f"{Vocabulary}is not exist")
        return False


def VocaRead(path='', sheet='Vocabulary'):
    df = pd.read_excel(path, sheet_name=sheet)
    df.to_csv('VocaTemp.csv', index=False)

    ddf = dd.read_csv('VocaTemp.csv')
    filtered_ddf = ddf[ddf['AlreadyProcessed'] == 0]
    vocabulary_list = filtered_ddf['Vocabulary'].compute().tolist()
    id_list = filtered_ddf['Id'].compute().tolist()
    return vocabulary_list, id_list


def VocaUpdateAlreadyPro(path='', sheet='Vocabulary'):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=3).value == 0:
            sheet.cell(row=row, column=3, value=1)
    workbook.save(path)


## Sentence
def SentUpdate(TestTime, path='', sheet='Sentence'):
    # 知道上次最後一行
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]
    last_row = sheet.max_row
    LastId = last_row + 1

    # 找出需要用GPT處理的LIST
    VocoList, VocoIdList = VocaRead(path=path)
    # print(VocoList, VocoIdList)

    for num, word in enumerate(VocoList):
        # print(word)
        print("Call GPT API")
        SentDf = GPT.GetSentenceGPT(word)

        # # for debug
        # data = {
        #     'Sentence': ['The grand opening', 'She was nervous', 'The job opening'],
        #     'Translation': ['Translation1', 'Translation2', 'Translation3'],
        #     'Usage': ['Usage1', 'Usage2', 'Usage3']
        # }
        # SentDf = pd.DataFrame(data)
        # print(type(SentDf))
        # print(SentDf)

        SentDf['Id'] = range(LastId, LastId + len(SentDf))
        SentDf = SentDf.assign(
            V_Id=VocoIdList[num],
            Level=1,
            TestDay=TestTime + 1,
            ErrorTime=0,
            FinishTime=0
        )

        # 重新排列列的顺序
        SentDf = SentDf[
            ['Id', 'V_Id', 'Sentence', 'Translation', 'Usage', 'Level', 'TestDay', 'ErrorTime', 'FinishTime']]
        # print(SentDf)

        # 將內容貼到Excel
        data_to_write = SentDf.values.tolist()
        # Write data to the specified rows
        for row_index, row_data in enumerate(data_to_write):
            for col_index, value in enumerate(row_data):
                sheet.cell(row=LastId + row_index, column=col_index + 1, value=value)

        # 更新Id
        LastId = LastId + len(SentDf)

    # 儲存Excel
    workbook.save(path)

    # 更新單字表示已處理
    VocaUpdateAlreadyPro(path=path)


def SentGetTest(TestTime, TestNum=40, path='',
                sheet='Sentence'):
    df = pd.read_excel(path, sheet_name=sheet)
    df.to_csv('SentTemp.csv', index=False)

    ddf = dd.read_csv('SentTemp.csv')
    filtered_ddf = ddf[ddf['TestDay'] <= TestTime]
    filtered_df = filtered_ddf.compute()
    filtered_df['RandomId'] = np.random.randint(1, 101, size=len(filtered_df))
    filtered_df.sort_values(by='RandomId', inplace=True)
    filtered_df.drop(columns=['RandomId'], inplace=True)

    # 只抓前幾個Row
    New_df = filtered_df.head(TestNum)
    New_df['Correct'] = 1
    New_df = New_df.reset_index(drop=True)

    # 處理音檔
    TTSPath = TTS.GetVoicePath()
    New_df.apply(lambda row: SentGetVoice(row['Id'], row['Level'], row['Sentence'], TTSPath), axis=1)

    return New_df


def SentGetVoice(Id, Level, Sentence, path):
    # 確認是否需要重新抓聲音
    match Level:
        case 1:
            GetNewVoice = False
        case 2:
            GetNewVoice = False
        case 3:
            GetNewVoice = SentChangeVoice(0.3)
        case 4:
            GetNewVoice = SentChangeVoice(0.8)
        case 5:
            GetNewVoice = SentChangeVoice(0.5)
        case 6:
            GetNewVoice = SentChangeVoice(0.7)
        case 7:
            GetNewVoice = True
    # 生成聲音
    if GetNewVoice == True:
        # 產生新的音訊檔
        voice_name = TTS.Azure_VoiceName(DeleteList=[])
        TTS.Azure_TTS(path, voice_name=voice_name, text=Sentence, GetFile=True, FileName=f'{Id}_voice')
        # print("New")
    else:
        # 先確認路徑是否存在
        CheckPath = f'{path}/{Id}_voice.wav'
        if TTS.Check_wav(CheckPath) == False:
            # 產生新的音訊檔
            voice_name = TTS.Azure_VoiceName(DeleteList=[])
            TTS.Azure_TTS(path, voice_name=voice_name, text=Sentence, GetFile=True, FileName=f'{Id}_voice')
            # print("Not Exist")
        else:
            # print("Exist")
            pass

def SentChangeVoice(probability):
    random_number = np.random.random()
    return random_number < probability

def SentChangeCorrect(MainDf, Id):
    MainDf.loc[MainDf['Id'] == Id, 'Correct'] = 0
    return MainDf


def SentUpdateAftTest(TestTime, MainDf, path='',
                      sheet='Sentence'):
    MainDf = MainDf.apply(lambda row: SentUpdateAftTestProcessRow(row, TestTime), axis=1)
    MainDf.drop(columns=['Correct'], inplace=True)

    # 將內容貼到Excel
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]

    data_to_write = MainDf.values.tolist()
    # Write data to the specified rows
    for row_data in data_to_write:
        for col_index, value in enumerate(row_data):
            sheet.cell(row=row_data[0], column=col_index + 1, value=value)
    # 儲存Excel
    workbook.save(path)

    data_dict = {'TestTime': TestTime}
    CoverUpdate(data_dict, path=path)


def SentUpdateAftTestProcessRow(row, TestTime):
    if row['Correct'] == 1:
        if row['Level'] + 1 > 7:
            row['FinishTime'] = row['FinishTime'] + 1
        else:
            row['Level'] = row['Level'] + 1
    elif row['Correct'] == 0:
        row['ErrorTime'] = row['ErrorTime'] + 1
        if row['Level'] - 2 < 1:
            row['Level'] = 1
        else:
            row['Level'] = row['Level'] - 2

    Level = row['Level']
    match Level:
        case 1:
            row['TestDay'] = TestTime + 1
        case 2:
            row['TestDay'] = TestTime + 1
        case 3:
            row['TestDay'] = TestTime + np.random.randint(1, 4)
        case 4:
            row['TestDay'] = TestTime + np.random.randint(3, 8)
        case 5:
            row['TestDay'] = TestTime + np.random.randint(6, 16)
        case 6:
            row['TestDay'] = TestTime + np.random.randint(12, 24)
        case 7:
            row['TestDay'] = TestTime + np.random.randint(30 * (row['FinishTime'] + 1),
                                                          60 * (row['FinishTime'] + 1) + 1)

    return row


if __name__ == '__main__':
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)

    # Get Excel Path
    # path = GetExcelPath()

